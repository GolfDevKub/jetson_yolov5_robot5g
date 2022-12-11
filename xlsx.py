from openpyxl import load_workbook, Workbook
from datetime import datetime

'''d = datetime.today()
time = d.strftime("%H.%M.%S")
date = d.strftime("%d-%m-%Y")'''
#print(type(n))
'''
def create_name_column():

    wb = Workbook()
    sheet = wb.active

    sheet['A1'] = "Date"
    sheet['B1'] = "Amount"    
    sheet['D1'] = "Time"    
    sheet['E1'] = "Total"

    wb.save("data_person.xlsx")    
'''
def excel_people(date, count, time):

    wb = load_workbook('data_people.xlsx')
    sheet = wb.active

    data = [
        (date, count, time)
    ]
    #print(type(data[0]))
    for row in data:
        sheet.append(row)
    wb.save('data_people.xlsx')


def excel_mask(date, count, time):

    wb = load_workbook('data_mask.xlsx')
    sheet = wb.active

    data = [
        (date, count, time)
    ]
    #print(type(data[0]))
    for row in data:
        sheet.append(row)
    wb.save('data_mask.xlsx')


def excel_nomask(date, count, time):

    wb = load_workbook('data_nomask.xlsx')
    sheet = wb.active

    data = [
        (date, count, time)
    ]
    #print(type(data[0]))
    for row in data:
        sheet.append(row)
    wb.save('data_nomask.xlsx')


def excel_upload_text(date, count, time, name_file):

    wb = load_workbook(name_file)
    sheet = wb.active

    data = [
        (date, count, time)
    ]
    #print(type(data[0]))
    for row in data:
        sheet.append(row)
    wb.save(name_file)  

#create_name_column()
#excel_upload_text(10, 10, 10,'data_nomask.xlsx')    
